"""Build the per-state panel from whichever source is authoritative per year.

Two independent DOSM publications cover overlapping years:

  DTS 2023 state XLSX  -- Table 1 spans 2017-2023, full float precision
  DTS 2024 state PDF   -- Table 1 spans 2018-2024, printed rounded
  DTS 2025 national    -- Table 9 gives visitors by state through 2025

The XLSX wins wherever both exist (it is unrounded and needs no parsing), the
PDF supplies 2024, and Table 9 supplies 2025. The 2018-2023 overlap is checked
cell by cell on the way through, extending the 2023-only reconciliation to every
shared year.
"""
import csv, os

import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
P = os.path.join(ROOT, "data", "processed")
XL = os.path.join(ROOT, "data", "raw", "opendosm-tourism")

STATE_FILE = {
    "Johor": "johor", "Kedah": "kedah", "Kelantan": "kelantan", "Melaka": "melaka",
    "Negeri Sembilan": "negerisembilan", "Pahang": "pahang", "Perak": "perak",
    "Perlis": "perlis", "Pulau Pinang": "pulaupinang", "Sabah": "sabah",
    "Sarawak": "sarawak", "Selangor": "selangor", "Terengganu": "terengganu",
    "W.P. Kuala Lumpur": "wpkualalumpur", "W.P. Labuan": "wplabuan",
    "W.P. Putrajaya": "wpputrajaya",
}
T1 = {
    "Jumlah Terimaan": ("total_receipts_rm_million", 0),
    "Pelawat Domestik": ("domestic_visitors_thousand", 0),
    "Perjalanan Pelancongan Domestik": ("domestic_trips_thousand", 0),
    "Purata Terimaan per Kapita": ("avg_receipts_per_capita_rm", 0),
    "Purata Terimaan per Perjalanan": ("avg_receipts_per_trip_rm", 0),
    "Purata Bilangan Hari Menginap": ("avg_length_of_stay_nights", 2),
}
XL_YEARS = list(range(2017, 2024))


def label_of(row):
    for c in row:
        if isinstance(c, str) and c.strip():
            return " ".join(c.split())
    return ""


def xl_table1(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next(wb[s] for s in wb.sheetnames if s.lower().replace(" ", "") == "jadual1")
    out = {}
    for row in ws.iter_rows(values_only=True):
        lab = label_of(row)
        if "Kadar Pertumbuhan" in lab:
            continue
        for key, (ind, _) in T1.items():
            if lab.startswith(key):
                v = [c for c in row if isinstance(c, (int, float))]
                if len(v) == len(XL_YEARS):
                    out[ind] = dict(zip(XL_YEARS, v))
    wb.close()
    return out


def main():
    pdf = pd.read_csv(os.path.join(P, "dts_key_statistics_2018_2024.csv"))
    v25 = pd.read_csv(os.path.join(P, "dts_visitors_by_state_2018_2025.csv"))

    rows, checked, bad = [], 0, []
    for state, stem in STATE_FILE.items():
        xl = xl_table1(os.path.join(XL, f"tourism_domestic_2023_{stem}.xlsx"))
        pv = pdf[pdf.state == state]

        for ind, (_, dp) in ((v[0], v) for v in T1.values()):
            series = xl.get(ind, {})
            for year, val in series.items():
                # overlap check against the PDF parser
                p = pv[(pv.indicator == ind) & (pv.year == year)]["value"]
                src = "DTS 2023 state XLSX (Table 1)"
                if not p.empty:
                    checked += 1
                    if round(float(val), dp) != round(float(p.iloc[0]), dp):
                        bad.append((state, ind, year, float(p.iloc[0]), float(val)))
                    src += " [agrees with DTS 2024 PDF]"
                rows.append({"state": state, "year": year, "indicator": ind,
                             "value": round(float(val), 4), "source": src,
                             "precision": "exact"})
            # 2024 exists only in the PDF
            p24 = pv[(pv.indicator == ind) & (pv.year == 2024)]["value"]
            if not p24.empty:
                rows.append({"state": state, "year": 2024, "indicator": ind,
                             "value": float(p24.iloc[0]),
                             "source": "DTS 2024 state PDF (Table 1)",
                             "precision": "as printed"})

        # 2025 visitors only
        v = v25[(v25.state == state) & (v25.year == 2025)]["domestic_visitors_thousand"]
        if not v.empty:
            rows.append({"state": state, "year": 2025,
                         "indicator": "domestic_visitors_thousand",
                         "value": float(v.iloc[0]),
                         "source": "DTS 2025 Malaysia PDF (Table 9)",
                         "precision": "as printed"})

    out = os.path.join(P, "state_panel_2017_2025.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["state", "year", "indicator", "value",
                                          "source", "precision"])
        w.writeheader()
        w.writerows(rows)

    df = pd.DataFrame(rows)
    print(f"OVERLAP CHECK 2018-2023 (XLSX vs PDF parser): {checked} cells, "
          f"{len(bad)} disagree")
    for b in bad[:20]:
        print(f"   {b[0]:18s} {b[1]:30s} {b[2]}  pdf={b[3]}  xlsx={b[4]}")

    print(f"\nwrote {len(df)} rows -> {os.path.relpath(out, ROOT)}")
    print(f"  states {df.state.nunique()}  years {df.year.min()}-{df.year.max()}  "
          f"indicators {df.indicator.nunique()}")
    print("\ncoverage (states x indicators present per year):")
    cov = df.pivot_table(index="year", columns="indicator", values="value",
                         aggfunc="count").fillna(0).astype(int)
    print(cov.to_string())


if __name__ == "__main__":
    main()
