"""Validate the PDF parser against an independent machine-readable publication.

Our 2023 per-state figures were read by our own regex parser out of the DTS 2024
state PDFs (which carry a 2018-2024 series and 2023-vs-2024 comparisons). DOSM
also published those same 2023 figures separately, as XLSX, in the DTS 2023 state
edition on storage.dosm.gov.my.

Comparing the two tests the parser itself against a clean source, rather than
testing it only for internal consistency. If they agree, the same parser's 2024
output -- which the whole dashboard sits on -- is trustworthy. If they disagree,
there is a bug affecting 2024 too.

The XLSX carries full float precision; the PDF prints rounded values. So each
comparison rounds the XLSX value to the PDF's printed precision before testing,
and any residual difference is a real disagreement, not a display artefact.
"""
import csv, os, sys

import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
P = os.path.join(ROOT, "data", "processed")
XL = os.path.join(ROOT, "data", "raw", "opendosm-tourism")

STATE_FILE = {
    "Johor": "johor", "Kedah": "kedah", "Kelantan": "kelantan", "Melaka": "melaka",
    "Negeri Sembilan": "negerisembilan", "Pahang": "pahang", "Perak": "perak",
    "Perlis": "perlis", "Pulau Pinang": "pulaupinang", "Sabah": "sabah",
    "Sarawak": "sarawak", "Selangor": "selangor", "Terengganu": "terengganu",
    "W.P. Kuala Lumpur": "wpkualalumpur", "W.P. Labuan": "wplabuan",
    "W.P. Putrajaya": "wpputrajaya",
}

# Table 1 row label -> (our indicator name, decimal places the PDF prints)
T1 = {
    "Jumlah Terimaan": ("total_receipts_rm_million", 0),
    "Pelawat Domestik": ("domestic_visitors_thousand", 0),
    "Perjalanan Pelancongan Domestik": ("domestic_trips_thousand", 0),
    "Purata Terimaan per Kapita": ("avg_receipts_per_capita_rm", 0),
    "Purata Terimaan per Perjalanan": ("avg_receipts_per_trip_rm", 0),
    "Purata Bilangan Hari Menginap": ("avg_length_of_stay_nights", 2),
}
T7 = {
    "A. Perbelanjaan oleh pelawat": "expenditure_by_visitors",
    "Membeli-belah": "shopping",
    "Pembelian bahan api kenderaan": "automotive_fuel",
    "Pengangkutan": "transportation",
    "Makanan & minuman": "food_and_beverage",
    "Penginapan": "accommodation",
    "Perbelanjaan sebelum perjalanan": "pre_trip_package_entrance_tickets",
    "Aktiviti-aktiviti lain": "other_activities",
    "B. Perbelanjaan oleh isi rumah": "expenditure_by_visited_households",
    "Jumlah Terimaan (A+B)": "total_receipts",
}
INCOME = {"≤ 1,000": "≤RM1,000", "1,001 - 3,000": "RM1,001-3,000",
          "3,001 - 5,000": "RM3,001-5,000", "5,001 - 10,000": "RM5,001-10,000",
          "≥ 10,001": "≥RM10,001"}
ACC = {"Rumah saudara": "relatives_friends_house", "Hotel/": "hotel",
       "Chalet/": "chalet", "Apartmen": "apartment",
       "Inap desa": "homestay_vacation_home", "Rumah rehat": "rest_house",
       "Tapak perkhemahan": "campsite", "Lain-lain": "other_accommodation"}


def label_of(row):
    for c in row:
        if isinstance(c, str) and c.strip():
            return " ".join(c.split())
    return ""


def nums(row):
    return [c for c in row if isinstance(c, (int, float))]


def xl_table1(ws):
    """Return {indicator: 2023 value}. Growth-rate rows have one fewer column,
    which is how they are told apart from value rows."""
    out = {}
    for row in ws.iter_rows(values_only=True):
        lab = label_of(row)
        if "Kadar Pertumbuhan" in lab:
            continue
        for key, (ind, _) in T1.items():
            if lab.startswith(key):
                v = nums(row)
                if len(v) == 7:          # 2017..2023
                    out[ind] = v[-1]
    return out


def xl_table7(ws):
    out = {}
    for row in ws.iter_rows(values_only=True):
        lab = label_of(row)
        for key, comp in T7.items():
            if lab.startswith(key):
                v = nums(row)
                if len(v) == 4:          # 2022, 2023 receipts; 2022, 2023 shares
                    out[comp] = (v[1], v[3])
    return out


def xl_shares(ws, mapping, stop_after=None):
    """Two-column percentage tables (2022, 2023)."""
    out, seen_stop = {}, False
    for row in ws.iter_rows(values_only=True):
        lab = label_of(row)
        if stop_after and stop_after in lab:
            seen_stop = True
        if stop_after and seen_stop and stop_after not in lab:
            break
        for key, name in mapping.items():
            if lab.startswith(key) and name not in out:
                v = nums(row)
                if len(v) == 2:
                    out[name] = v[1]
    return out


def close(a, b, dp):
    """PDF prints to `dp` decimals; agree if the XLSX rounds to the same string."""
    if a is None or b is None:
        return False
    return round(float(b), dp) == round(float(a), dp)


def main():
    k = pd.read_csv(os.path.join(P, "dts_key_statistics_2018_2024.csv"))
    k = k[k.year == 2023]
    r7 = pd.read_csv(os.path.join(P, "dts_receipts_by_component_2023_2024.csv"))
    inc = pd.read_csv(os.path.join(P, "dts_visitor_income_class.csv"))
    acc = pd.read_csv(os.path.join(P, "dts_accommodation_type.csv"))

    rows, mismatches = [], []
    for state, stem in STATE_FILE.items():
        path = os.path.join(XL, f"tourism_domestic_2023_{stem}.xlsx")
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = {s.lower().replace(" ", ""): s for s in wb.sheetnames}

        def sheet(*cands):
            for c in cands:
                if c in sheets:
                    return wb[sheets[c]]
            return None

        # --- Table 1 ---
        t1 = xl_table1(sheet("jadual1"))
        for ind, (_, dp) in ((v[0], v) for v in T1.values()):
            ours = k[(k.state == state) & (k.indicator == ind)]["value"]
            ours = None if ours.empty else float(ours.iloc[0])
            theirs = t1.get(ind)
            ok = close(ours, theirs, dp)
            rows.append([state, "table_1", ind, ours, theirs, dp, ok])
            if not ok:
                mismatches.append((state, "table_1", ind, ours, theirs))

        # --- Table 7: receipts (RM '000, printed whole) and shares (1 dp) ---
        t7 = xl_table7(sheet("jadual7"))
        for comp, (recv, share) in t7.items():
            row = r7[(r7.state == state) & (r7.component == comp)]
            if row.empty:
                mismatches.append((state, "table_7", comp, None, recv))
                continue
            o_r = row["receipts_rm_thousand_2023"].iloc[0]
            o_s = row["share_pct_2023"].iloc[0]
            for tag, ours, theirs, dp in (("receipts", o_r, recv, 0),
                                          ("share", o_s, share, 1)):
                ok = close(ours, theirs, dp)
                rows.append([state, "table_7", f"{comp}:{tag}", ours, theirs, dp, ok])
                if not ok:
                    mismatches.append((state, "table_7", f"{comp}:{tag}", ours, theirs))

        # --- Table 13 income class ---
        xi = xl_shares(sheet("jadual13a", "jadual13"), INCOME)
        for cls, theirs in xi.items():
            row = inc[(inc.state == state) & (inc.income_class == cls)]
            ours = None if row.empty else float(row["share_pct_2023"].iloc[0])
            ok = close(ours, theirs, 1)
            rows.append([state, "table_13_income", cls, ours, theirs, 1, ok])
            if not ok:
                mismatches.append((state, "table_13_income", cls, ours, theirs))

        # --- Table 12 accommodation type ---
        xa = xl_shares(sheet("jadual11&12", "jadual12"), ACC, stop_after=None)
        for name, theirs in xa.items():
            row = acc[(acc.state == state) & (acc.accommodation_type == name)]
            ours = None if row.empty else float(row["share_pct_2023"].iloc[0])
            ok = close(ours, theirs, 1)
            rows.append([state, "table_12_accom", name, ours, theirs, 1, ok])
            if not ok:
                mismatches.append((state, "table_12_accom", name, ours, theirs))
        wb.close()

    out = os.path.join(P, "reconciliation_2023_pdf_vs_xlsx.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["state", "table", "field", "pdf_parser_value",
                    "xlsx_published_value", "compared_at_dp", "agree"])
        w.writerows(rows)

    n = len(rows)
    bad = sum(1 for r in rows if not r[6])
    print(f"compared {n} cells across 16 states -> {os.path.relpath(out, ROOT)}")
    print(f"  agree    : {n - bad}")
    print(f"  disagree : {bad}")
    by = {}
    for r in rows:
        t = r[1]
        by.setdefault(t, [0, 0])
        by[t][0] += 1
        by[t][1] += (not r[6])
    print("\n  by table:")
    for t, (tot, b) in sorted(by.items()):
        print(f"    {t:18s} {tot-b:4d}/{tot:4d} agree" + ("" if not b else f"   <-- {b} DISAGREE"))
    if mismatches:
        print(f"\n  first 30 disagreements:")
        for m in mismatches[:30]:
            print(f"    {m[0]:18s} {m[1]:16s} {m[2]:38s} pdf={m[3]}  xlsx={m[4]}")


if __name__ == "__main__":
    main()
